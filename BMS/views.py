from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import user_passes_test
from .models import *
import datetime
from django.core import validators
from openpyxl import load_workbook
#model 
from django.contrib.auth.models import User
from AS.models import Account,StudentInfo,Package
# Create your views here.

#is superuser?
@user_passes_test(lambda u: u.is_superuser,login_url='/AS/login/')
def stuinfo_import(request):
    messege = ''
    if request.method == 'POST':
        myfile = request.FILES.get('myfile')
        validate=myfile.name.split('.')
        #確認副檔名為excel檔
        if validate[-1]=='xlsx' or validate[-1]=='xls':
            now = str(datetime.datetime.now().year)+'-'+str(datetime.datetime.now().month)+'-'+str(datetime.datetime.now().day)\
                +' '+str(datetime.datetime.now().hour)+'_'+str(datetime.datetime.now().minute)+'_'+str(datetime.datetime.now().second)
            name=now+' '+myfile.name
            with open('stuinfo excel/'+name,'wb') as fp:
                for chunk in myfile.chunks():
                    fp.write(chunk)
            messege="上傳成功! "+stuinfo_insert(name)
        else :
            messege="請選擇excel檔案上傳"
    return render(request, 'stuinfo_import.html', locals())

#mail_import
@user_passes_test(lambda u: u.is_superuser,login_url='/AS/login/')
def mail_import(request):
    if request.method == 'POST':
        myfile = request.FILES.get('myfile')
        validate=myfile.name.split('.')
        #確認副檔名為excel檔
        if validate[-1]=='xlsx' or validate[-1]=='xls':
            now = str(datetime.datetime.now().year)+'-'+str(datetime.datetime.now().month)+'-'+str(datetime.datetime.now().day)\
                +' '+str(datetime.datetime.now().hour)+'_'+str(datetime.datetime.now().minute)+'_'+str(datetime.datetime.now().second)
            name=now+' '+myfile.name
            with open('mail excel/'+name,'wb') as fp:
                for chunk in myfile.chunks():
                    fp.write(chunk)
            messege="上傳成功! "+mail_insert(name)
        else :
            messege="請選擇excel檔案上傳"
    return render(request, 'mail_import.html', locals())
    
#開檔寫入資料庫 
#表格順序為Account,First-name,Last-name,Gender,Departmane,Grade
def stuinfo_insert(filename):
    column = {
        1:'Account',
        2:'First-name',
        3:'Last-name',
        4:'Gender',
        5:'Department',
        6:'Grade',
    }
    #try:
    wb = load_workbook('stuinfo excel//'+filename,'read_only')
    sheet = wb.active
    maxrow=sheet.max_row
    maxcol=sheet.max_column
    #確認表格順序
    for i in range(1,maxcol+1):
        if sheet.cell(row=1, column=i).value != column[i] :
            messege='請將資料表第一列依照 Account,First-name,Last-name,Gender,Department,Grade 順序排列'
            return messege
    #type check
    for c in range(2,maxrow+1):
        if len(sheet.cell(row=c, column=1).value) != 8:
            messege = '第'+str(c)+'橫排的學號格式錯誤(8位數含英文)'
            return messege
        if sheet.cell(row=c, column=4).value != 'M' and sheet.cell(row=c, column=4).value != 'F':
            messege = '第'+str(c)+'橫排的性別格式錯誤只能為(M或F)'
            return messege
        if  sheet.cell(row=c, column=6).value <1 or  sheet.cell(row=c, column=6).value >4:
            messege = '第'+str(c)+'橫排的年級格式錯誤只能為(1~4)'
            return messege
    #新增資料
    for c in range(2,maxrow+1):
        user = User.objects.create_user( username=sheet.cell(row=c, column=1).value , first_name=sheet.cell(row=c, column=2).value,\
        last_name=sheet.cell(row=c, column=3).value , password='ggininder30cm')
        user.save()
        account = Account(user=user,permission=3)
        account.save()
        stuinfo = StudentInfo(account=account,gender=sheet.cell(row=c, column=4).value,\
            grade=sheet.cell(row=c, column=6).value,department=sheet.cell(row=c, column=5).value,room='x',bed=0)
        stuinfo.save()
    messege='新增成功'
    return messege

def mail_insert(filename):
    column = {
        1:'Category', #Mail or Package or Promt
        2:'Sender',
        3:'Receiver',
    }
    #try:
    wb = load_workbook('mail excel//'+filename,'read_only')
    sheet = wb.active
    maxrow=sheet.max_row
    maxcol=sheet.max_column
    #確認表格順序
    for i in range(1,maxcol+1):
        if sheet.cell(row=1, column=i).value != column[i] :
            messege='請將資料表第一列依照 Category Sender Receiver 順序排列'
            return messege
    #type check
    for c in range(2,maxrow+1):
        if sheet.cell(row=c, column=1).value != 'Mail' and \
            sheet.cell(row=c, column=1).value != 'Package' and \
             sheet.cell(row=c, column=1).value != 'Prompt':
            messege = '第'+str(c)+'橫排的內容錯誤( Mail 或 Package 或 Prompt )'
            return messege
        try:
            user=User.objects.get(username=sheet.cell(row=c, column=3).value)
        except:
            messege = '第'+str(c)+'橫排的內容錯誤( 不存在該名學生 )'
            return messege
    for c in range(2,maxrow+1):
        mail=Package(category=sheet.cell(row=c, column=1).value,\
         sender=sheet.cell(row=c, column=2).value,receiver=Account.objects.get(user=user))
        mail.save()
    messege='新增成功'
    return messege